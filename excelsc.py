import openpyxl 
import json
  
# Give the location of the file 
path = "E:\\home\\excelscript\\All_Divident_chart.xlsx"
  
# To open the workbook  
# workbook object is created 
wb_obj = openpyxl.load_workbook(path,data_only=True) 
  
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = None
sheetname = None
singlesheet = None
#sheet_obj = wb_obj.active 
print(wb_obj.sheetnames)
sheetname = wb_obj.sheetnames
singlesheet = wb_obj[sheetname[17]]
sheet_obj = wb_obj[sheetname[17]]
  
# Cell objects also have a row, column,  
# and coordinate attributes that provide 
# location information for the cell. 
  
# Note: The first row or  
# column integer is 1, not 0. 
  
# Cell object is created by using  
# sheet object's cell() method. 
cell_obj = None
cell_obj = sheet_obj.cell(row = 1, column = 1) 


data = {}
data['product_name'] = wb_obj.sheetnames[17]
data['total_months'] = sheet_obj.cell(row = 24, column = 1).value
data['total_amount_of_payment'] = sheet_obj.cell(row = 25, column = 2).value
data['total_divident'] = sheet_obj.cell(row = 25, column = 3).value
listmonth = []
max_col = sheet_obj.max_column 
print(max_col)
max_row = sheet_obj.max_row
print(max_row)
print(sheet_obj)
for i in range(5, sheet_obj.max_row+1):
    index = 1
    monthdata = {};
    print([cell.value for cell in sheet_obj[i]])
    #print(sheet_obj[i])
    row = None
    row = [cell.value for cell in sheet_obj[i]] # sheet[n] gives nth row (list of cells)
    monthdata['month'] = row[0]
    monthdata['amount_of_payment'] = row[1]
    monthdata['divident'] = row[2]
    monthdata['paid'] = row[3]
    monthdata['comparison'] = row[4]
    monthdata['int_rate_bank'] = row[5]
    monthdata['int_rate_kopuram'] = row[6]
    listmonth.append(monthdata)
    
data['monthly_emi'] = listmonth

with open(wb_obj.sheetnames[17]+'.json', 'w') as json_file:
    json.dump(data, json_file)
